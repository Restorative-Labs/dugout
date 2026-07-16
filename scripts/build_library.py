"""Convert DugOut_Tip_Library_Seed_v0.2.xlsx -> dugout/js/library.js

The spreadsheet is the source of truth (DUG-37). This script is the only thing that
should ever write library.js, so a corrected sheet can be re-converted verbatim.
"""
import json
import re

import openpyxl

SRC = "/Users/derelljenkins/Library/CloudStorage/OneDrive-Personal/Restorative Labs LLC/04 — Dugout/DugOut_Tip_Library_Seed_v0.2.xlsx"
OUT = "/Users/derelljenkins/Library/CloudStorage/OneDrive-Personal/Restorative Labs LLC/04 — Dugout/dugout/js/library.js"

CHECKPOINT_KEY = {
    "Stance": "stance",
    "Load": "load",
    "Stride": "stride",
    "Hips Fire": "hips",
    "Contact": "contact",
    "Finish": "finish",
}

# Which faults the pose engine can actually measure, and with what detector.
# Established by probing the real footage — see docs/POSE-PROBE.md.
# null => not measurable offline (needs ball flight, bat tracking, or detail pose lacks).
DETECTOR = {
    "ST-01": "stance.baseWidth",
    "ST-02": None,   # needs home plate reference
    "ST-03": None,   # grip detail below pose resolution
    "ST-04": "stance.weightBias",
    "LD-01": "load.noLoad",
    "LD-02": "load.drift",
    "LD-03": None,   # bat wrap — bat not tracked
    "LD-04": None,   # hitch — wrist dip too marginal to call
    "SR-01": "stride.early",
    "SR-02": "stride.overstride",
    "SR-03": "stride.bucket",
    "SR-04": None,   # foot angle — MoveNet has ankle but no toe keypoint
    "SR-05": "stride.late",
    "SR-06": "stride.headMove",
    "HF-01": "hips.rotationSpeed",
    "HF-02": "hips.separation",
    "HF-03": "hips.flyOpen",
    "HF-04": "hips.backsideCollapse",
    "CT-01": None,   # ground-ball contact — needs ball flight
    "CT-02": None,   # push pattern — needs pitch location
    "CT-03": None,   # rolling over — needs ball flight
    "CT-04": None,   # attack angle — needs bat tracking
    "CT-05": "contact.casting",
    "CT-06": "contact.batDrag",
    "CT-07": "contact.headPull",
    "FN-01": None,   # barrel decel — needs bat tracking
    "FN-02": "finish.balance",
    "FN-03": "finish.height",
    "FN-04": None,   # style toggle, not a fault (see TOGGLES)
}

# FN-04 is explicitly "a PHILOSOPHY TOGGLE in the app, not a fault entry" per the sheet.
TOGGLES = {"FN-04"}

WEIGHTED_RE = re.compile(r"overload|underload|weighted|plyo", re.I)
BANDS = [("7–10", "7-10"), ("11–13", "11-13"), ("14–18", "14-18")]
ALL_BANDS = ["7-10", "11-13", "14-18"]


def split_list(v, seps=";·"):
    if not v or str(v).strip() in ("—", "-", ""):
        return []
    s = str(v).replace("\n", " ")
    parts = re.split("[" + seps + "]", s)
    return [p.strip().strip("'").strip() for p in parts if p.strip() and p.strip() != "—"]


def parse_bands(v):
    """'All (emphasize 7-10)' -> all three; '11-13, 14-18' -> those two."""
    s = str(v or "")
    head = s.split("(")[0]
    if "All" in head:
        bands = list(ALL_BANDS)
    else:
        bands = [key for dash, key in BANDS if dash in head]
        if not bands:
            bands = list(ALL_BANDS)
    emph = []
    tail = s[s.find("(") :] if "(" in s else ""
    for dash, key in BANDS:
        if dash in tail and re.search(r"emphasi|great", tail, re.I):
            emph.append(key)
    return bands, emph


def drill_rung(name):
    n = name.lower()
    if "live" in n or "machine" in n:
        return "live"
    if "front toss" in n:
        return "frontToss"
    if "soft toss" in n or "softtoss" in n:
        return "softToss"
    return "tee"  # default: closed, lowest rung


def build_tips(wb):
    ws = wb["Tip Library"]
    rows = list(ws.iter_rows(values_only=True))
    tips = []
    for r in rows[1:]:
        if not r[0]:
            continue
        tid = str(r[0]).strip()
        bands, emph = parse_bands(r[8])
        drills = []
        for d in split_list(r[5]):
            drills.append({"name": d, "rung": drill_rung(d), "weighted": bool(WEIGHTED_RE.search(d))})
        tips.append(
            {
                "id": tid,
                "checkpoint": CHECKPOINT_KEY[str(r[1]).strip()],
                "fault": str(r[2]).strip(),
                "detection": str(r[3]).strip(),
                "tip": str(r[4]).strip(),
                "drills": drills,
                "cues": [c for part in split_raw(r[6], "·") for c in quoted_phrases(part)],
                # raw text keeps the trainer-facing nuance ("...as a rigid rule");
                # avoidPhrases is what the red-line filter actually matches on.
                "cuesToAvoid": split_list(r[7], seps=";·"),
                "avoidPhrases": [p.lower() for p in quoted_phrases(r[7], quoted_only=True)],
                "ageBands": bands,
                "emphasize": emph,
                "schoolTags": split_list(r[9], seps=";+"),
                "tier": str(r[10]).strip()[0] if r[10] else "C",
                "sources": str(r[11] or "").strip(),
                "notes": str(r[12] or "").strip(),
                "detector": DETECTOR.get(tid),
                "isToggle": tid in TOGGLES,
            }
        )
    return tips


# A closing quote is an apostrophe NOT followed by a word character, so "can't" stays
# intact while "'Turn fast' (outcome cue)" still closes correctly.
QUOTED = re.compile(r"'((?:[^']|'(?=\w))*)'")


def split_raw(v, seps="·"):
    """Split on separators WITHOUT stripping quotes.

    quoted_phrases() must see the original quoting to find phrase boundaries. Stripping a
    leading quote first makes the apostrophe in "shortstop's" look like the opening quote,
    which silently yields "s left ear".
    """
    if not v or str(v).strip() in ("—", "-", ""):
        return []
    s = str(v).replace("\n", " ")
    return [p.strip() for p in re.split("[" + seps + "]", s) if p.strip() and p.strip() != "—"]


def quoted_phrases(v, quoted_only=False):
    """"'Swing down' / 'chop down at the ball'" -> two separate phrases.

    Cells pack several quoted cues plus parenthetical glosses into one string. Cues get
    read aloud to a kid, so they must come out clean; the red-line filter likewise needs
    each phrase separately or it will never match real text. Unquoted cells (e.g. LD-01's
    "Internal body-part checklists during live reps") fall back to the whole string.
    """
    s = str(v or "").strip()
    found = [p.strip() for p in QUOTED.findall(s) if p.strip()]
    if found:
        return found
    if quoted_only:
        return []  # unquoted prose is a note, not a cue phrase — nothing to match on
    return [s] if s and s != "—" else []


def build_glossary(wb):
    ws = wb["Cue Glossary"]
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        status = str(r[2]).strip()
        out.append(
            {
                "id": str(r[0]).strip(),
                "phrases": quoted_phrases(r[1]),
                "phrase": (quoted_phrases(r[1]) or [""])[0],
                "status": "deprecated" if status.startswith("DEPRECATED") else ("contested" if status.startswith("CONTESTED") else "active"),
                "statusText": status,
                "lineage": str(r[3] or "").strip(),
                "why": str(r[4] or "").strip(),
                "replacement": str(r[5] or "").strip(),
                "tier": str(r[6]).strip()[0] if r[6] else "C",
            }
        )
    return out


def build_schools(wb):
    ws = wb["Schools of Thought"]
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        ids = re.findall(r"[A-Z]{2}-\d{2}", str(r[4] or ""))
        out.append(
            {
                "preset": str(r[0]).strip(),
                "lineage": str(r[1] or "").strip(),
                "beliefs": str(r[2] or "").strip(),
                "signatureCues": str(r[3] or "").strip(),
                "emphasisIds": ids,
                "tier": str(r[5]).strip()[0] if r[5] else "C",
                "notes": str(r[6] or "").strip(),
            }
        )
    return out


def build_rules(wb):
    ws = wb["Engine Rules"]
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        out.append(
            {
                "id": str(r[0]).strip(),
                "rule": str(r[1]).strip(),
                "meaning": str(r[2]).strip(),
                "tier": str(r[3]).strip()[0] if r[3] else "C",
            }
        )
    return out


wb = openpyxl.load_workbook(SRC)
tips = build_tips(wb)
glossary = build_glossary(wb)
schools = build_schools(wb)
rules = build_rules(wb)

deprecated = sorted({p.lower() for g in glossary if g["status"] == "deprecated" for p in g["phrases"]})

js = f"""// GENERATED FILE — do not edit by hand.
// Source of truth: DugOut_Tip_Library_Seed_v0.2.xlsx (DUG-37).
// Regenerate with scripts/build_library.py after the trainer revises the sheet.
//
// {len(tips)} tip entries · {len(glossary)} cue-glossary entries · {len(schools)} school presets · {len(rules)} engine rules
// `detector: null` means the fault is not measurable offline (needs ball flight, bat
// tracking, or detail pose lacks) — see docs/POSE-PROBE.md. Those faults are never
// guessed at; the checkpoint is graded on the entries that can be measured.

window.DUGOUT_LIBRARY = {{
  tips: {json.dumps(tips, indent=2, ensure_ascii=False)},

  cueGlossary: {json.dumps(glossary, indent=2, ensure_ascii=False)},

  schools: {json.dumps(schools, indent=2, ensure_ascii=False)},

  engineRules: {json.dumps(rules, indent=2, ensure_ascii=False)},

  // Phrases that must never reach output in any rendering path (CG-01, CG-02).
  // They exist here only as red-line data.
  deprecatedCues: {json.dumps(deprecated, indent=2, ensure_ascii=False)}
}};
"""

import os

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w") as f:
    f.write(js)

# ---- report ----
print("tips: %d" % len(tips))
by_cp = {}
for t in tips:
    by_cp.setdefault(t["checkpoint"], []).append(t)
for cp in ["stance", "load", "stride", "hips", "contact", "finish"]:
    ts = by_cp[cp]
    det = [t["id"] for t in ts if t["detector"]]
    nod = [t["id"] for t in ts if not t["detector"] and not t["isToggle"]]
    print("  %-8s %d entries | detectable: %-28s | gap: %s" % (cp, len(ts), ",".join(det) or "-", ",".join(nod) or "-"))
print("toggles: %s" % [t["id"] for t in tips if t["isToggle"]])
print("weighted drills: %s" % sorted({t["id"] for t in tips for d in t["drills"] if d["weighted"]}))
print("deprecated cues: %s" % deprecated)
print("wrote %s (%d bytes)" % (OUT, os.path.getsize(OUT)))
